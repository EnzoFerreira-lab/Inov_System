"""
Importador do arquivo real de DRE da contabilidade (o mesmo arquivo que eles já usam
todo mês, com uma aba por obra: "OBRA 251", "OBRA 265", "DEMAIS OBRAS", etc.).

Não assume posição fixa de linha/coluna (isso varia de aba pra aba, cada obra tem um
histórico de tamanho diferente). Em vez disso, localiza a estrutura pelo texto:

    "Receitas Brutas"  -> início da seção de receita (linha de cabeçalho de mês fica
                           uma linha acima)
    "...Total..."       -> fim de uma seção (linha de total, não é uma categoria)
    "Custos"            -> início da seção de custo
    "...Total..."       -> fim da seção de custo

As colunas de mês são identificadas pelo próprio texto do cabeçalho
("Janeiro/2026", "Fevereiro/2026"...) — colunas de acumulado histórico
("Março a Dez/23", "Jan a Dez/24"...) são ignoradas automaticamente, pois não
batem com esse padrão.
"""

import re
import unicodedata
import difflib
import datetime

import openpyxl

MESES_PT = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

PADRAO_MES_ANO = re.compile(
    r"^\s*(janeiro|fevereiro|mar[cç]o|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)\s*[/\s]\s*(\d{2,4})\s*$",
    re.IGNORECASE,
)

# Colunas de período agregado sem detalhamento mensal, ex: "Out a Dez/2025",
# "Março a Dez/23", "Jan a Dez/24". Não são um mês específico — são um bloco
# histórico anterior ao início do controle mês a mês daquela obra.
PADRAO_PERIODO_AGREGADO = re.compile(
    r"^\s*\S+\s+a\s+\S+\s*/\s*\d{2,4}\s*$",
    re.IGNORECASE,
)

SHEETS_IGNORADAS = ("TOTAIS",)  # consolidado redundante — o sistema já calcula isso ao vivo em /totais


def _sem_acento(texto):
    if texto is None:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(texto))
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _normaliza(texto):
    return re.sub(r"\s+", " ", _sem_acento(texto).lower()).strip()


def _extrair_codigo_e_nome(label):
    """'Alimentação - 320' -> ('320', 'Alimentação'); 'Ferias' -> (None, 'Ferias')."""
    label = re.sub(r"\s+", " ", str(label)).strip()
    m = re.search(r"-\s*(\d+)\s*$", label)
    codigo = None
    if m:
        codigo = m.group(1)
        label = label[: m.start()].strip(" -")
    return codigo, label


def _ano_completo(ano):
    ano = int(ano)
    return ano + 2000 if ano < 100 else ano


def _mapear_colunas_de_mes(ws, header_row, col_max=40):
    """Retorna {coluna_idx: (ano, mes)} para as colunas cujo cabeçalho é 'Mês/Ano'."""
    colunas = {}
    for c in range(2, col_max + 1):
        valor = ws.cell(row=header_row, column=c).value
        if not valor:
            continue
        m = PADRAO_MES_ANO.match(str(valor))
        if m:
            mes = MESES_PT[_sem_acento(m.group(1)).lower()]
            ano = _ano_completo(m.group(2))
            colunas[c] = (ano, mes)
    return colunas


def _mapear_colunas_historicas(ws, header_row, col_max=40):
    """Retorna {coluna_idx: 'Out a Dez/2025'} para colunas de período agregado
    (sem detalhamento mensal), distintas das colunas de mês único."""
    colunas = {}
    for c in range(2, col_max + 1):
        valor = ws.cell(row=header_row, column=c).value
        if not valor:
            continue
        texto = str(valor).strip()
        if PADRAO_MES_ANO.match(texto):
            continue
        if PADRAO_PERIODO_AGREGADO.match(texto):
            colunas[c] = texto
    return colunas


def _localizar_secoes(ws, max_linhas=80, col_rotulo=1):
    """
    Varre a coluna de rótulo e devolve (receitas_row, custos_row) — as linhas onde
    aparecem "Receitas Brutas"/"Receita Bruta" e "Custos".
    """
    receitas_row = None
    custos_row = None
    for r in range(1, max_linhas):
        valor = ws.cell(row=r, column=col_rotulo).value
        if not valor:
            continue
        texto = _normaliza(valor)
        if "receita" in texto and "brut" in texto and receitas_row is None:
            receitas_row = r
        elif texto == "custos" and receitas_row is not None and custos_row is None:
            custos_row = r
            break
    return receitas_row, custos_row


def _linhas_de_categoria(ws, linha_inicio, max_linhas=80, col_rotulo=1):
    """A partir de linha_inicio+1, devolve linhas de categoria até achar uma linha
    de 'total' (que encerra a seção) ou uma linha vazia."""
    linhas = []
    r = linha_inicio + 1
    while r < linha_inicio + max_linhas:
        valor = ws.cell(row=r, column=col_rotulo).value
        if valor is None or not str(valor).strip():
            r += 1
            if r - linha_inicio > 3 and not linhas:
                break
            continue
        texto = _normaliza(valor)
        if "total" in texto:
            break
        linhas.append(r)
        r += 1
    return linhas, r  # r = linha do total (ou onde parou)


def _obra_codigo_e_nome_do_sheet(ws, sheet_name):
    nome_upper = sheet_name.strip().upper()

    m = re.search(r"OBRA\s*(\d+)", nome_upper)
    if m:
        codigo = m.group(1)
    elif "DEMAIS OBRAS" in nome_upper:
        codigo = "DEMAIS"
    else:
        return None, None, None

    status = "em_andamento"
    nome_obra = sheet_name.strip()

    for r in range(1, 10):
        valor = ws.cell(row=r, column=1).value or ws.cell(row=r, column=2).value
        if not valor:
            continue
        texto = str(valor).strip()
        texto_norm = _normaliza(texto)
        if texto_norm == "obra encerrada":
            status = "encerrada"
        elif texto_norm == "obra em andamento":
            status = "em_andamento"
        elif texto_norm.startswith("obra") and len(texto) > len("obra encerrada") + 2:
            nome_obra = texto

    return codigo, nome_obra, status


def _obter_ou_criar_obra(cur, empresa_id, codigo, nome, status):
    cur.execute("SELECT id FROM obras WHERE codigo = ?", (codigo,))
    row = cur.fetchone()
    if row:
        return row["id"], False

    cur.execute(
        "INSERT INTO obras (empresa_id, nome, codigo, status) VALUES (?, ?, ?, ?)",
        (empresa_id, nome, codigo, status),
    )
    return cur.lastrowid, True


def _obter_ou_criar_categoria(cur, cache, codigo, nome, tipo):
    chave_codigo = codigo
    nome_norm = _normaliza(nome)

    if chave_codigo and chave_codigo in cache["por_codigo"]:
        return cache["por_codigo"][chave_codigo], False

    if nome_norm in cache["por_nome"]:
        return cache["por_nome"][nome_norm], False

    proximos = difflib.get_close_matches(nome_norm, cache["por_nome"].keys(), n=1, cutoff=0.85)
    if proximos:
        return cache["por_nome"][proximos[0]], False

    cur.execute("SELECT COALESCE(MAX(ordem), 0) + 1 AS prox FROM categorias_conta")
    ordem = cur.fetchone()["prox"]
    cur.execute(
        "INSERT INTO categorias_conta (codigo, nome, tipo, ordem, ativo) VALUES (?, ?, ?, ?, 1)",
        (codigo, nome, tipo, ordem),
    )
    novo_id = cur.lastrowid
    cache["por_nome"][nome_norm] = novo_id
    if codigo:
        cache["por_codigo"][codigo] = novo_id
    return novo_id, True


def _carregar_cache_categorias(cur):
    cur.execute("SELECT id, codigo, nome FROM categorias_conta")
    linhas = cur.fetchall()
    return {
        "por_codigo": {r["codigo"]: r["id"] for r in linhas if r["codigo"]},
        "por_nome": {_normaliza(r["nome"]): r["id"] for r in linhas},
    }


PADRAO_TITULO_MES_ANO = re.compile(
    r"(janeiro|fevereiro|mar[cç]o|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)\s*/\s*(\d{2,4})",
    re.IGNORECASE,
)


class RegistroImportacao:
    """
    Intermedia toda gravação feita por uma importação.

    Existe por dois motivos:

    1. **Não perder correção manual.** Um valor ajustado à mão na tela de
       Lançar Dados fica com origem='manual'. Reimportar o mesmo mês
       sobrescrevia esse ajuste em silêncio. Agora, por padrão, o valor manual
       é preservado e o conflito é reportado — quem importa decide.

    2. **Poder desfazer.** Antes de alterar uma célula, o estado anterior é
       gravado em importacao_itens, o que permite reverter um arquivo enviado
       por engano sem restaurar o banco inteiro.

    Sem importacao_id (usado nos testes), grava normalmente mas não registra
    histórico.
    """

    LIMITE_CONFLITOS = 25  # amostra suficiente para a tela; não guarda os 17 mil

    def __init__(self, cur, importacao_id=None, sobrescrever_manuais=False):
        self.cur = cur
        self.importacao_id = importacao_id
        self.sobrescrever_manuais = sobrescrever_manuais

        self.lancamentos_gravados = 0
        self.saldos_gravados = 0
        self.manuais_preservados = 0
        self.manuais_sobrescritos = 0
        self.conflitos = []

    def _registrar_item(self, tabela, obra_id, categoria_id, anterior,
                        mes=None, ano=None, periodo=None):
        if self.importacao_id is None:
            return

        self.cur.execute(
            """
            INSERT INTO importacao_itens
                (importacao_id, tabela, obra_id, categoria_id, mes, ano,
                 periodo_descricao, existia, valor_anterior, origem_anterior)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (self.importacao_id, tabela, obra_id, categoria_id, mes, ano, periodo,
             1 if anterior else 0,
             anterior["valor"] if anterior else None,
             anterior["origem"] if anterior else None),
        )

    def gravar_lancamento(self, obra_id, categoria_id, mes, ano, valor, agora):
        """Grava um valor mensal. Devolve False se preservou um lançamento manual."""
        self.cur.execute(
            "SELECT valor, origem FROM lancamentos "
            "WHERE obra_id = ? AND categoria_id = ? AND mes = ? AND ano = ?",
            (obra_id, categoria_id, mes, ano),
        )
        anterior = self.cur.fetchone()

        if anterior and anterior["origem"] == "manual":
            if not self.sobrescrever_manuais:
                self.manuais_preservados += 1
                if len(self.conflitos) < self.LIMITE_CONFLITOS:
                    self.conflitos.append({
                        "obra_id": obra_id, "categoria_id": categoria_id,
                        "mes": mes, "ano": ano,
                        "valor_manual": anterior["valor"], "valor_planilha": valor,
                    })
                return False
            self.manuais_sobrescritos += 1

        self._registrar_item("lancamentos", obra_id, categoria_id, anterior, mes=mes, ano=ano)

        self.cur.execute(
            """
            INSERT INTO lancamentos (obra_id, categoria_id, mes, ano, valor, origem, atualizado_em)
            VALUES (?, ?, ?, ?, ?, 'importacao', ?)
            ON CONFLICT (obra_id, categoria_id, mes, ano)
            DO UPDATE SET valor = excluded.valor, origem = 'importacao',
                          atualizado_em = excluded.atualizado_em
            """,
            (obra_id, categoria_id, mes, ano, valor, agora),
        )
        self.lancamentos_gravados += 1
        return True

    def gravar_saldo_anterior(self, obra_id, categoria_id, periodo, valor, agora):
        """Grava um período histórico agregado (ex: 'Out a Dez/2025')."""
        self.cur.execute(
            "SELECT valor, origem FROM saldos_anteriores "
            "WHERE obra_id = ? AND categoria_id = ? AND periodo_descricao = ?",
            (obra_id, categoria_id, periodo),
        )
        anterior = self.cur.fetchone()

        self._registrar_item("saldos_anteriores", obra_id, categoria_id, anterior, periodo=periodo)

        self.cur.execute(
            """
            INSERT INTO saldos_anteriores
                (obra_id, categoria_id, periodo_descricao, valor, origem, atualizado_em)
            VALUES (?, ?, ?, ?, 'importacao', ?)
            ON CONFLICT (obra_id, categoria_id, periodo_descricao)
            DO UPDATE SET valor = excluded.valor, atualizado_em = excluded.atualizado_em
            """,
            (obra_id, categoria_id, periodo, valor, agora),
        )
        self.saldos_gravados += 1
        return True


def desfazer_importacao(cur, importacao_id):
    """
    Reverte uma importação, devolvendo cada célula ao valor que tinha antes.
    Linhas que a importação criou do zero são apagadas.

    Só faz sentido desfazer a importação mais recente ainda ativa — desfazer uma
    antiga por cima de outra mais nova ressuscitaria valores obsoletos. Quem
    chama é responsável por essa checagem (ver a rota em app.py).
    """
    cur.execute("SELECT * FROM importacoes WHERE id = ?", (importacao_id,))
    importacao = cur.fetchone()
    if not importacao:
        raise ValueError("Importação não encontrada.")
    if importacao["desfeita_em"]:
        raise ValueError("Essa importação já foi desfeita.")

    cur.execute("SELECT * FROM importacao_itens WHERE importacao_id = ?", (importacao_id,))
    itens = cur.fetchall()

    restaurados = 0
    removidos = 0

    for item in itens:
        if item["tabela"] == "lancamentos":
            onde = "obra_id = ? AND categoria_id = ? AND mes = ? AND ano = ?"
            chave = (item["obra_id"], item["categoria_id"], item["mes"], item["ano"])
        else:
            onde = "obra_id = ? AND categoria_id = ? AND periodo_descricao = ?"
            chave = (item["obra_id"], item["categoria_id"], item["periodo_descricao"])

        if item["existia"]:
            cur.execute(
                f"UPDATE {item['tabela']} SET valor = ?, origem = ? WHERE {onde}",
                (item["valor_anterior"], item["origem_anterior"] or "importacao") + chave,
            )
            restaurados += 1
        else:
            cur.execute(f"DELETE FROM {item['tabela']} WHERE {onde}", chave)
            removidos += 1

    cur.execute(
        "UPDATE importacoes SET desfeita_em = ? WHERE id = ?",
        (datetime.datetime.now().isoformat(), importacao_id),
    )

    return {"restaurados": restaurados, "removidos": removidos}


def importar_depto_tecnico(ws, cur, empresa_id_padrao, cache_categorias, resumo, registro):
    """
    Lê a aba 'DEPTO TÉCNICO' (formato diferente das obras: rótulo na coluna B,
    um único mês de valores na coluna C, sem as taxas de imposto — é despesa
    administrativa da empresa, não de uma obra específica).
    """
    ano_mes = None
    for r in range(1, 6):
        for c in (1, 2):
            valor = ws.cell(row=r, column=c).value
            if valor:
                m = PADRAO_TITULO_MES_ANO.search(str(valor))
                if m:
                    mes = MESES_PT[_sem_acento(m.group(1)).lower()]
                    ano = _ano_completo(m.group(2))
                    ano_mes = (ano, mes)
    if not ano_mes:
        return False
    ano, mes = ano_mes

    receitas_row, custos_row = _localizar_secoes(ws, col_rotulo=2)
    if not receitas_row or not custos_row:
        return False

    obra_id, criada = _obter_ou_criar_obra(cur, empresa_id_padrao, "DEPTO-TEC", "Departamento Técnico (Administrativo)", "em_andamento")
    if criada:
        resumo["obras_criadas"].append("DEPTO-TEC - Departamento Técnico (Administrativo)")

    linhas_receita, _ = _linhas_de_categoria(ws, receitas_row, col_rotulo=2)
    linhas_custo, _ = _linhas_de_categoria(ws, custos_row, col_rotulo=2)

    agora = datetime.datetime.now().isoformat()

    for linhas, tipo in ((linhas_receita, "receita"), (linhas_custo, "custo")):
        for r in linhas:
            label_bruto = ws.cell(row=r, column=2).value
            codigo_cat, nome_cat = _extrair_codigo_e_nome(label_bruto)
            if not nome_cat:
                continue

            categoria_id, criada_cat = _obter_ou_criar_categoria(cur, cache_categorias, codigo_cat, nome_cat, tipo)
            if criada_cat:
                resumo["categorias_criadas"].append(f"{nome_cat} ({tipo})")

            valor_bruto = ws.cell(row=r, column=3).value
            if valor_bruto in (None, ""):
                continue
            try:
                valor_bruto = float(valor_bruto)
            except (TypeError, ValueError):
                continue

            valor = valor_bruto if tipo == "receita" else -valor_bruto

            if registro.gravar_lancamento(obra_id, categoria_id, mes, ano, valor, agora):
                resumo["lancamentos_gravados"] += 1

    return True


def importar_planilha_dre_real(caminho_arquivo, cur, empresa_id_padrao, registro=None):
    """
    Lê o arquivo real de DRE (multi-abas) e grava obras/categorias/lançamentos.
    Retorna um dicionário com o resumo da importação.

    'registro' é um RegistroImportacao, que preserva lançamentos manuais e
    guarda o estado anterior para permitir desfazer. Se não vier um, cria um
    sem histórico (comportamento útil em teste e em scripts).
    """
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)

    if registro is None:
        registro = RegistroImportacao(cur)

    cache_categorias = _carregar_cache_categorias(cur)

    resumo = {
        "abas_processadas": [],
        "abas_ignoradas": [],
        "obras_criadas": [],
        "categorias_criadas": [],
        "lancamentos_gravados": 0,
        "saldos_anteriores_gravados": 0,
    }

    agora = datetime.datetime.now().isoformat()

    for sheet_name in wb.sheetnames:
        nome_upper = sheet_name.strip().upper()

        if any(nome_upper.startswith(prefixo) or prefixo in nome_upper for prefixo in SHEETS_IGNORADAS):
            resumo["abas_ignoradas"].append((sheet_name, "consolidado redundante — o sistema calcula isso ao vivo em /totais"))
            continue

        if "DEPTO" in nome_upper or "DEPARTAMENTO" in nome_upper:
            ws = wb[sheet_name]
            sucesso = importar_depto_tecnico(ws, cur, empresa_id_padrao, cache_categorias, resumo, registro)
            if sucesso:
                resumo["abas_processadas"].append(sheet_name)
            else:
                resumo["abas_ignoradas"].append((sheet_name, "não consegui identificar mês/seções nessa aba"))
            continue

        ws = wb[sheet_name]

        codigo_obra, nome_obra, status = _obra_codigo_e_nome_do_sheet(ws, sheet_name)
        if not codigo_obra:
            resumo["abas_ignoradas"].append((sheet_name, "não parece ser uma aba de obra"))
            continue

        receitas_row, custos_row = _localizar_secoes(ws)
        if not receitas_row or not custos_row:
            resumo["abas_ignoradas"].append((sheet_name, "não encontrei as seções 'Receitas Brutas'/'Custos'"))
            continue

        header_row = receitas_row - 1
        colunas_mes = _mapear_colunas_de_mes(ws, header_row)
        colunas_historicas = _mapear_colunas_historicas(ws, header_row)
        if not colunas_mes:
            resumo["abas_ignoradas"].append((sheet_name, "não encontrei colunas de mês no cabeçalho"))
            continue

        obra_id, criada = _obter_ou_criar_obra(cur, empresa_id_padrao, codigo_obra, nome_obra, status)
        if criada:
            resumo["obras_criadas"].append(f"{codigo_obra} - {nome_obra}")

        linhas_receita, _ = _linhas_de_categoria(ws, receitas_row)
        linhas_custo, _ = _linhas_de_categoria(ws, custos_row)

        for linhas, tipo in ((linhas_receita, "receita"), (linhas_custo, "custo")):
            for r in linhas:
                label_bruto = ws.cell(row=r, column=1).value
                codigo_cat, nome_cat = _extrair_codigo_e_nome(label_bruto)
                if not nome_cat:
                    continue

                categoria_id, criada_cat = _obter_ou_criar_categoria(
                    cur, cache_categorias, codigo_cat, nome_cat, tipo
                )
                if criada_cat:
                    resumo["categorias_criadas"].append(f"{nome_cat} ({tipo})")

                for col_idx, (ano, mes) in colunas_mes.items():
                    valor_bruto = ws.cell(row=r, column=col_idx).value
                    if valor_bruto in (None, ""):
                        continue
                    try:
                        valor_bruto = float(valor_bruto)
                    except (TypeError, ValueError):
                        continue

                    # Na planilha, receita é positiva e custo é negativo (uma linha de
                    # custo positiva é um estorno/crédito, e deve REDUZIR o total, não
                    # somar). Convertendo pro nosso padrão (valor = magnitude, tipo diz
                    # o sinal): receita fica como está; custo inverte o sinal.
                    valor = valor_bruto if tipo == "receita" else -valor_bruto

                    if registro.gravar_lancamento(obra_id, categoria_id, mes, ano, valor, agora):
                        resumo["lancamentos_gravados"] += 1

                for col_idx, periodo_texto in colunas_historicas.items():
                    valor_bruto = ws.cell(row=r, column=col_idx).value
                    if valor_bruto in (None, ""):
                        continue
                    try:
                        valor_bruto = float(valor_bruto)
                    except (TypeError, ValueError):
                        continue
                    if valor_bruto == 0:
                        continue

                    valor = valor_bruto if tipo == "receita" else -valor_bruto

                    if registro.gravar_saldo_anterior(obra_id, categoria_id, periodo_texto, valor, agora):
                        resumo["saldos_anteriores_gravados"] += 1

        resumo["abas_processadas"].append(sheet_name)

    return resumo

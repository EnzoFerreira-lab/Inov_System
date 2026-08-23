"""
Geração do DRE em Excel para download — mesma estrutura da tela, formatada
pra impressão/envio por e-mail.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FONTE_PADRAO = "Arial"

FORMATO_MOEDA = '#,##0.00;(#,##0.00);"-"'

COR_CABECALHO = "1D4ED8"
COR_DESTAQUE = "EEF1F8"
COR_NEGATIVO = "B42318"


def _estilo_titulo(cell, tamanho=14, cor="FFFFFF", negrito=True, fundo=None):
    cell.font = Font(name=FONTE_PADRAO, size=tamanho, bold=negrito, color=cor)
    if fundo:
        cell.fill = PatternFill("solid", fgColor=fundo)


def _linha_valores(ws, linha, rotulo, valores, negrito=False, moeda=True, fundo=None, cor_texto=None):
    cell = ws.cell(row=linha, column=1, value=rotulo)
    cell.font = Font(name=FONTE_PADRAO, bold=negrito, color=cor_texto or "000000")
    if fundo:
        cell.fill = PatternFill("solid", fgColor=fundo)

    for i, v in enumerate(valores, start=2):
        c = ws.cell(row=linha, column=i, value=round(v, 2) if v is not None else None)
        c.font = Font(name=FONTE_PADRAO, bold=negrito, color=cor_texto or "000000")
        if moeda:
            c.number_format = FORMATO_MOEDA
        c.alignment = Alignment(horizontal="right")
        if fundo:
            c.fill = PatternFill("solid", fgColor=fundo)


def gerar_excel_dre_obra(obra, ano, meses, meses_nome, categorias, totais, acumulado,
                         periodos_historicos=None, totais_historicos=None, total_geral=None):
    """
    Monta o DRE da obra na mesma ordem de colunas da planilha de origem:
    períodos anteriores em bloco, meses do ano, acumulado do ano e total geral.
    """
    periodos_historicos = periodos_historicos or []
    totais_historicos = totais_historicos or {}
    total_geral = total_geral or acumulado

    wb = Workbook()
    ws = wb.active
    ws.title = f"DRE {obra['codigo']}"[:31]

    cabecalho = (
        ["Categoria"]
        + list(periodos_historicos)
        + [f"{meses_nome[m][:3]}/{a}" for a, m in meses]
        + [f"Acum. {ano}", "Total geral"]
    )
    ultima_coluna = get_column_letter(len(cabecalho))

    ws.merge_cells(f"A1:{ultima_coluna}1")
    ws["A1"] = obra["nome"]
    _estilo_titulo(ws["A1"], tamanho=14, fundo=COR_CABECALHO)
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{ultima_coluna}2")
    ws["A2"] = f"{obra['empresa_nome']} · Código {obra['codigo']} · DRE {ano}"
    ws["A2"].font = Font(name=FONTE_PADRAO, size=10, italic=True)

    linha = 4
    for i, texto in enumerate(cabecalho, start=1):
        c = ws.cell(row=linha, column=i, value=texto)
        c.font = Font(name=FONTE_PADRAO, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        c.alignment = Alignment(horizontal="center" if i > 1 else "left", wrap_text=True)
    linha += 1

    def valores_categoria(cat):
        historicos = [cat.get("historicos", {}).get(p, 0) for p in periodos_historicos]
        mensais = [cat["valores"].get((a, m), 0) for a, m in meses]
        return historicos + mensais + [sum(mensais), sum(mensais) + sum(historicos)]

    def valores_totais(campo):
        return (
            [totais_historicos[p][campo] for p in periodos_historicos]
            + [totais[(a, m)][campo] for a, m in meses]
            + [acumulado[campo], total_geral[campo]]
        )

    ws.cell(row=linha, column=1, value="Receitas Brutas").font = Font(name=FONTE_PADRAO, bold=True)
    linha += 1
    for cat in categorias:
        if cat["tipo"] != "receita":
            continue
        nome = cat["nome"] + (f" - {cat['codigo']}" if cat["codigo"] else "")
        _linha_valores(ws, linha, nome, valores_categoria(cat))
        linha += 1

    _linha_valores(ws, linha, "= Receitas Brutas Total", valores_totais("receita_total"),
                   negrito=True, fundo=COR_DESTAQUE)
    linha += 1

    ws.cell(row=linha, column=1, value="Custos").font = Font(name=FONTE_PADRAO, bold=True)
    linha += 1
    for cat in categorias:
        if cat["tipo"] != "custo":
            continue
        nome = cat["nome"] + (f" - {cat['codigo']}" if cat["codigo"] else "")
        _linha_valores(ws, linha, nome, [-v for v in valores_categoria(cat)], cor_texto=COR_NEGATIVO)
        linha += 1

    _linha_valores(ws, linha, "= Custos Total", [-v for v in valores_totais("custos_total")],
                   negrito=True, fundo=COR_DESTAQUE, cor_texto=COR_NEGATIVO)
    linha += 1

    _linha_valores(ws, linha, "= Lucro / Prejuízo Bruto", valores_totais("lucro_bruto"),
                   negrito=True, fundo=COR_DESTAQUE)
    linha += 1

    for chave, rotulo in [
        ("impostos_servicos", "(-) Impostos s/ Serviços"),
        ("irpj_csll", "(-) IRPJ e CSLL"),
        ("despesa_administrativa", "(-) Desp. Administrativas/Técnico"),
        ("despesa_financeira", "(-) Despesas Financeiras"),
    ]:
        _linha_valores(ws, linha, rotulo, [-v for v in valores_totais(chave)], cor_texto=COR_NEGATIVO)
        linha += 1

    _linha_valores(ws, linha, "LUCRO/PREJUÍZO LÍQUIDO DA OBRA", valores_totais("lucro_liquido"),
                   negrito=True, fundo="FFE8A3")
    linha += 1

    if periodos_historicos:
        linha += 1
        aviso = ws.cell(
            row=linha, column=1,
            value="As primeiras colunas são períodos que a planilha de origem traz "
                  "agregados, sem detalhamento mês a mês.",
        )
        aviso.font = Font(name=FONTE_PADRAO, size=9, italic=True, color="6B7280")

    ws.column_dimensions["A"].width = 34
    for i in range(2, len(cabecalho) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "B5"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def gerar_excel_dre_consolidado(ano, meses, meses_nome, totais, acumulado):
    wb = Workbook()
    ws = wb.active
    ws.title = "Totais Consolidado"

    ws.merge_cells("A1:N1")
    ws["A1"] = f"DRE Consolidado — Todas as Obras — {ano}"
    _estilo_titulo(ws["A1"], tamanho=14, fundo=COR_CABECALHO)
    ws.row_dimensions[1].height = 24

    linha = 3
    cabecalho = ["Linha"] + [f"{meses_nome[m][:3]}/{a}" for a, m in meses] + ["Acumulado"]
    for i, texto in enumerate(cabecalho, start=1):
        c = ws.cell(row=linha, column=i, value=texto)
        c.font = Font(name=FONTE_PADRAO, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        c.alignment = Alignment(horizontal="center" if i > 1 else "left")
    linha += 1

    vals_receita = [totais[(a, m)]["receita_total"] for a, m in meses] + [acumulado["receita_total"]]
    _linha_valores(ws, linha, "Receitas Brutas Total", vals_receita, negrito=True, fundo=COR_DESTAQUE)
    linha += 1

    vals_custo = [-totais[(a, m)]["custos_total"] for a, m in meses] + [-acumulado["custos_total"]]
    _linha_valores(ws, linha, "Custos Total", vals_custo, negrito=True, fundo=COR_DESTAQUE, cor_texto=COR_NEGATIVO)
    linha += 1

    vals_bruto = [totais[(a, m)]["lucro_bruto"] for a, m in meses] + [acumulado["lucro_bruto"]]
    _linha_valores(ws, linha, "= Lucro / Prejuízo Bruto", vals_bruto, negrito=True, fundo=COR_DESTAQUE)
    linha += 1

    for chave, rotulo in [
        ("impostos_servicos", "(-) Impostos s/ Serviços"),
        ("irpj_csll", "(-) IRPJ e CSLL"),
        ("despesa_administrativa", "(-) Desp. Administrativas/Técnico"),
        ("despesa_financeira", "(-) Despesas Financeiras"),
    ]:
        vals = [-totais[(a, m)][chave] for a, m in meses] + [-acumulado[chave]]
        _linha_valores(ws, linha, rotulo, vals, cor_texto=COR_NEGATIVO)
        linha += 1

    vals_liquido = [totais[(a, m)]["lucro_liquido"] for a, m in meses] + [acumulado["lucro_liquido"]]
    _linha_valores(ws, linha, "LUCRO/PREJUÍZO LÍQUIDO CONSOLIDADO", vals_liquido, negrito=True, fundo="FFE8A3")

    ws.column_dimensions["A"].width = 34
    for i in range(2, len(cabecalho) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.freeze_panes = "B4"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

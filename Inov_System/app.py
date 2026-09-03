from flask import (
    Flask, render_template, request, redirect, url_for, session, flash,
    send_file, abort,
)
import os
import re
import secrets
import logging
import sqlite3
import datetime
import pandas as pd
import openpyxl
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash

from db import conectar, criar_tabelas
from dre import (
    calcular_dre_obra,
    calcular_dre_consolidado,
    CAMPOS_TOTAIS,
)
from dre_import import importar_planilha_dre_real, RegistroImportacao, desfazer_importacao
from contimatic import buscar_partidas
import backup
from dre_export import gerar_excel_dre_obra, gerar_excel_dre_consolidado

app = Flask(__name__)

# A chave assina o cookie de sessão. Um valor fixo no código permitiria a
# qualquer um forjar uma sessão de administrador, então em produção ela vem do
# ambiente. Sem a variável, sorteia uma chave nova a cada inicialização: é
# seguro, com o efeito colateral aceitável de derrubar as sessões no restart.
app.secret_key = os.environ.get("INOV_SECRET_KEY")
if not app.secret_key:
    app.secret_key = secrets.token_hex(32)
    logging.getLogger(__name__).warning(
        "INOV_SECRET_KEY não definida — usando chave temporária. "
        "Defina a variável de ambiente em produção para manter as sessões entre reinícios."
    )

UPLOAD_FOLDER = "uploads"
EXPORT_FOLDER = "exports"
ALLOWED_EXTENSIONS = {"xlsx", "xls"}

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["EXPORT_FOLDER"] = EXPORT_FOLDER

MESES_NOME = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

MESES_ABREV = [
    "", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
    "Jul", "Ago", "Set", "Out", "Nov", "Dez",
]


# ---------------------------------------------------------------------------
# Formatação — tudo que aparece na tela usa o padrão brasileiro (1.234,56).
# O Python formata como 1,234.56, então trocamos os separadores de posição.
# ---------------------------------------------------------------------------

@app.template_filter("moeda")
def filtro_moeda(valor, casas=2):
    """1234.5 -> '1.234,50'. Sem o 'R$', que fica a cargo do template."""
    try:
        valor = float(valor or 0)
    except (TypeError, ValueError):
        return "0,00"

    texto = f"{abs(valor):,.{casas}f}"
    texto = texto.replace(",", "\x00").replace(".", ",").replace("\x00", ".")
    return f"-{texto}" if valor < 0 else texto


@app.template_filter("moeda_curta")
def filtro_moeda_curta(valor):
    """Versão compacta para cartões de indicador: 1482310.55 -> '1,48 mi'."""
    try:
        valor = float(valor or 0)
    except (TypeError, ValueError):
        return "0,00"

    sinal = "-" if valor < 0 else ""
    absoluto = abs(valor)

    if absoluto >= 1_000_000:
        return f"{sinal}{filtro_moeda(absoluto / 1_000_000)} mi"
    if absoluto >= 1_000:
        return f"{sinal}{filtro_moeda(absoluto / 1_000, 1)} mil"
    return f"{sinal}{filtro_moeda(absoluto)}"


@app.template_filter("pct")
def filtro_percentual(valor, casas=1):
    """0.1315 já em pontos percentuais (13.15) -> '13,2%'."""
    try:
        valor = float(valor or 0)
    except (TypeError, ValueError):
        return "0,0%"
    return f"{filtro_moeda(valor, casas)}%"


@app.template_filter("data_br")
def filtro_data_br(valor):
    """'2026-06-01' -> '01/06/2026'. Devolve o original se não for uma data ISO."""
    if not valor:
        return "—"
    partes = str(valor)[:10].split("-")
    if len(partes) != 3:
        return str(valor)
    ano, mes, dia = partes
    return f"{dia}/{mes}/{ano}"


@app.template_filter("competencia_br")
def filtro_competencia_br(valor):
    """'2026-06' -> 'Jun/2026' (usado nas vigências de taxa)."""
    if not valor:
        return "—"
    partes = str(valor).split("-")
    if len(partes) < 2:
        return str(valor)
    try:
        return f"{MESES_ABREV[int(partes[1])]}/{partes[0]}"
    except (ValueError, IndexError):
        return str(valor)


app.jinja_env.globals.update(
    meses_nome=MESES_NOME,
    meses_abrev=MESES_ABREV,
)


# ---------------------------------------------------------------------------
# Proteção CSRF
#
# Sem isso, uma página em outro site conseguiria disparar um POST no sistema
# usando a sessão de quem estivesse logado — e há POSTs destrutivos aqui
# (excluir obra apaga todos os lançamentos dela). Cada formulário carrega um
# token ligado à sessão, e todo método que altera dado é obrigado a apresentá-lo.
# ---------------------------------------------------------------------------

METODOS_QUE_ALTERAM = ("POST", "PUT", "PATCH", "DELETE")


def token_csrf():
    if "_csrf" not in session:
        session["_csrf"] = secrets.token_urlsafe(32)
    return session["_csrf"]


@app.before_request
def validar_csrf():
    if request.method not in METODOS_QUE_ALTERAM:
        return

    esperado = session.get("_csrf")
    enviado = request.form.get("_csrf") or request.headers.get("X-CSRF-Token", "")

    if not esperado or not secrets.compare_digest(str(enviado), str(esperado)):
        abort(400, "Sessão expirada ou requisição inválida. Recarregue a página e tente de novo.")


@app.context_processor
def injetar_csrf():
    from markupsafe import Markup

    def campo_csrf():
        """Insere o campo oculto do token. Todo <form method="POST"> precisa dele."""
        return Markup(f'<input type="hidden" name="_csrf" value="{token_csrf()}">')

    return {"campo_csrf": campo_csrf}


# ---------------------------------------------------------------------------
# Consultas auxiliares compartilhadas por várias telas
# ---------------------------------------------------------------------------

def anos_com_dados(cur, ano_atual):
    """
    Anos que realmente têm movimento, para o seletor não oferecer uma janela
    fixa (que mostrava anos vazios e escondia anos com dado fora dela).

    O 'valor <> 0' importa: a planilha traz colunas de meses futuros ainda em
    branco, e o importador cria a linha zerada. Sem o filtro, o seletor
    ofereceria anos que abrem um DRE inteiramente vazio.
    """
    cur.execute("SELECT DISTINCT ano FROM lancamentos WHERE valor <> 0 ORDER BY ano")
    anos = {r["ano"] for r in cur.fetchall()}
    anos.add(ano_atual)
    anos.add(datetime.date.today().year)
    return sorted(anos)


# ---------------------------------------------------------------------------
# Permissões
#
# O corte não é por obra: todo funcionário enxerga e opera todas as obras —
# importa, exporta, lança e atualiza. O que fica reservado à dona da empresa
# são as ações que destroem dado ou mudam a regra do jogo para todo mundo:
# excluir empresa ou obra, alterar as taxas do DRE e administrar usuários.
# ---------------------------------------------------------------------------

def usuario_eh_admin():
    return session.get("usuario_papel") == "admin"


def exigir_admin():
    """Usar depois de exigir_login(). Devolve resposta de bloqueio ou None."""
    if not usuario_logado():
        return redirect(url_for("login"))
    if not usuario_eh_admin():
        flash(
            "Essa ação é restrita à administradora do sistema. "
            "Peça para ela, ou solicite que seu usuário seja promovido.",
            "erro",
        )
        return redirect(url_for("dashboard"))
    return None


@app.context_processor
def injetar_permissoes():
    """Deixa 'eh_admin' disponível em todo template, para esconder o que não pode."""
    return {"eh_admin": usuario_eh_admin()}


def anos_da_obra(cur, obra_id):
    """
    Anos com movimento **desta** obra. O seletor era global, então oferecia
    2024/2025/2026 para todas — e 27 das 53 obras abriam num ano em que não
    têm lançamento mensal nenhum, parecendo que o sistema não tinha os dados.
    """
    cur.execute(
        "SELECT DISTINCT ano FROM lancamentos WHERE obra_id = ? AND valor <> 0 ORDER BY ano",
        (obra_id,),
    )
    return [r["ano"] for r in cur.fetchall()]


def escolher_ano_da_obra(cur, obra_id, ano_pedido):
    """
    Decide qual ano abrir. Sem pedido explícito, usa o ano corrente se a obra
    tiver dados nele; senão, o ano mais recente que ela tem. Obras antigas
    (encerradas há anos) abriam num DRE vazio por causa disso.
    """
    anos = anos_da_obra(cur, obra_id)
    ano_corrente = datetime.date.today().year

    if ano_pedido:
        ano = ano_pedido
    elif ano_corrente in anos:
        ano = ano_corrente
    elif anos:
        ano = anos[-1]
    else:
        ano = ano_corrente

    if ano not in anos:
        anos = sorted(set(anos) | {ano})

    return ano, anos


def montar_comparativo(atual, anterior, ano_anterior):
    """
    Variação de cada linha do DRE contra o ano anterior, como a planilha mostra.

    A variação percentual é omitida quando o ano anterior foi zero: dividir por
    zero não dá "aumento de 100%", dá uma informação que não existe. Nesses
    casos a tela mostra só o valor absoluto.
    """
    linhas = []
    for campo, rotulo in [
        ("receita_total", "Receita Bruta"),
        ("custos_total", "Custos"),
        ("lucro_bruto", "Lucro Bruto"),
        ("lucro_liquido", "Lucro Líquido"),
    ]:
        valor_atual = atual.get(campo, 0) or 0
        valor_anterior = anterior.get(campo, 0) or 0
        diferenca = valor_atual - valor_anterior

        linhas.append({
            "campo": campo,
            "rotulo": rotulo,
            "atual": valor_atual,
            "anterior": valor_anterior,
            "diferenca": diferenca,
            "percentual": (diferenca / abs(valor_anterior) * 100) if valor_anterior else None,
            # Em custo, subir é ruim; nas demais linhas, subir é bom.
            "subir_e_bom": campo != "custos_total",
        })

    return {"ano_anterior": ano_anterior, "linhas": linhas}


def listar_empresas(cur):
    cur.execute("SELECT id, nome FROM empresas ORDER BY nome")
    return [dict(r) for r in cur.fetchall()]


def obra_ids_da_empresa(cur, empresa_id=None):
    """
    Ids das obras a consolidar. Sem empresa escolhida, todas.

    Existe porque /totais e /dashboard somavam indistintamente as obras de
    todas as empresas — com um único cliente ninguém notava, mas o segundo
    cliente tornaria o consolidado errado sem dar nenhum erro.
    """
    if empresa_id:
        cur.execute("SELECT id FROM obras WHERE empresa_id = ?", (empresa_id,))
    else:
        cur.execute("SELECT id FROM obras")
    return [r["id"] for r in cur.fetchall()]


def criar_pastas():
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(EXPORT_FOLDER, exist_ok=True)


def preparar_ambiente():
    """
    Garante pastas e schema na inicialização.

    Roda ao importar o módulo, não só em `python app.py`: servido por WSGI
    (gunicorn, waitress) o bloco __main__ nunca executa, e o sistema subia
    sem as tabelas. criar_tabelas() usa CREATE TABLE IF NOT EXISTS e só semeia
    o que ainda não existe, então repetir é inofensivo.
    """
    criar_pastas()
    criar_tabelas()

    # Uma cópia por dia, no primeiro acesso do dia. Não derruba o sistema se
    # falhar: ficar sem backup é ruim, mas não subir é pior.
    try:
        destino = backup.backup_diario()
        if destino:
            app.logger.info("Backup diário do banco gerado em %s", destino)
    except Exception:
        app.logger.exception("Não foi possível gerar o backup diário do banco")


preparar_ambiente()


def usuario_logado():
    return "usuario_id" in session


def exigir_login():
    if not usuario_logado():
        return redirect(url_for("login"))
    return None


def parse_valor_br(texto):
    """
    Converte o que o usuário digitou num float, aceitando os formatos que
    aparecem na prática: '1234.56', '1234,56', '1.234,56' e 'R$ 1.234,56'.
    Devolve 0.0 quando o campo está vazio ou ilegível — um valor inválido não
    pode derrubar o salvamento da grade inteira.
    """
    if texto is None:
        return 0.0

    limpo = re.sub(r"[^\d,.\-]", "", str(texto)).strip()
    if not limpo:
        return 0.0

    # Quando aparecem os dois separadores, o último é o decimal.
    if "," in limpo and "." in limpo:
        if limpo.rfind(",") > limpo.rfind("."):
            limpo = limpo.replace(".", "").replace(",", ".")
        else:
            limpo = limpo.replace(",", "")
    elif "," in limpo:
        limpo = limpo.replace(",", ".")

    try:
        return abs(float(limpo))
    except ValueError:
        return 0.0


def erro_ao_salvar(mensagem_conflito, excecao):
    """
    Separa "o usuário tentou duplicar um registro" de "algo quebrou".

    Antes, qualquer falha virava a mesma mensagem sobre duplicidade, o que
    escondia o erro real e deixava o diagnóstico no chute. Agora só o conflito
    de unicidade recebe a mensagem amigável; o resto vai para o log do servidor.
    """
    if isinstance(excecao, sqlite3.IntegrityError):
        flash(mensagem_conflito, "erro")
    else:
        app.logger.exception("Falha inesperada ao gravar no banco")
        flash(
            "Erro inesperado ao salvar. O detalhe técnico foi registrado no log do servidor.",
            "erro",
        )


def arquivo_permitido(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def meses_do_ano(ano):
    return [(ano, m) for m in range(1, 13)]


# ---------------------------------------------------------------------------
# Autenticação
# ---------------------------------------------------------------------------

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email")
        senha = request.form.get("senha")

        conn = conectar()
        cur = conn.cursor()
        cur.execute("SELECT * FROM usuarios WHERE email = ?", (email,))
        usuario = cur.fetchone()
        conn.close()

        if usuario and check_password_hash(usuario["senha_hash"], senha or ""):
            if not usuario["ativo"]:
                flash("Esse acesso foi desativado. Fale com a administradora.", "erro")
                return render_template("login.html")

            # Recria a sessão no login: evita que um identificador de sessão
            # obtido antes da autenticação continue valendo depois dela.
            token = session.get("_csrf")
            session.clear()
            if token:
                session["_csrf"] = token

            session["usuario_id"] = usuario["id"]
            session["usuario_nome"] = usuario["nome"]
            session["usuario_papel"] = usuario["papel"]
            return redirect(url_for("dashboard"))

        flash("E-mail ou senha inválidos.", "erro")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/minha-conta", methods=["GET", "POST"])
def minha_conta():
    redir = exigir_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        atual = request.form.get("senha_atual") or ""
        nova = request.form.get("senha_nova") or ""
        confirmacao = request.form.get("senha_confirmacao") or ""

        cur.execute("SELECT senha_hash FROM usuarios WHERE id = ?", (session["usuario_id"],))
        usuario = cur.fetchone()

        if not usuario or not check_password_hash(usuario["senha_hash"], atual):
            flash("A senha atual está incorreta.", "erro")
        elif len(nova) < 6:
            flash("A nova senha precisa ter pelo menos 6 caracteres.", "erro")
        elif nova != confirmacao:
            flash("A confirmação não confere com a nova senha.", "erro")
        else:
            cur.execute(
                "UPDATE usuarios SET senha_hash = ? WHERE id = ?",
                (generate_password_hash(nova), session["usuario_id"]),
            )
            conn.commit()
            flash("Senha alterada com sucesso.", "sucesso")

    cur.execute("SELECT * FROM usuarios WHERE id = ?", (session["usuario_id"],))
    usuario = cur.fetchone()
    conn.close()

    return render_template("minha_conta.html", usuario=usuario)


# ---------------------------------------------------------------------------
# Usuários (só a administradora)
# ---------------------------------------------------------------------------

@app.route("/usuarios", methods=["GET", "POST"])
def usuarios():
    redir = exigir_admin()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        senha = request.form.get("senha") or ""
        papel = request.form.get("papel") if request.form.get("papel") in ("admin", "comum") else "comum"

        if not (nome and email and senha):
            flash("Preencha nome, e-mail e senha.", "erro")
        elif len(senha) < 6:
            flash("A senha precisa ter pelo menos 6 caracteres.", "erro")
        else:
            try:
                cur.execute(
                    """
                    INSERT INTO usuarios (nome, email, senha_hash, papel, ativo, criado_em)
                    VALUES (?, ?, ?, ?, 1, ?)
                    """,
                    (nome, email, generate_password_hash(senha), papel,
                     datetime.datetime.now().isoformat()),
                )
                conn.commit()
                flash(f"Usuário {nome} cadastrado.", "sucesso")
            except Exception as e:
                erro_ao_salvar("Já existe um usuário com esse e-mail.", e)

    cur.execute("SELECT * FROM usuarios ORDER BY ativo DESC, nome")
    lista = cur.fetchall()
    conn.close()

    return render_template("usuarios.html", usuarios=lista)


@app.route("/usuarios/<int:usuario_id>/papel", methods=["POST"])
def alterar_papel_usuario(usuario_id):
    redir = exigir_admin()
    if redir:
        return redir

    papel = request.form.get("papel")
    if papel not in ("admin", "comum"):
        flash("Papel inválido.", "erro")
        return redirect(url_for("usuarios"))

    conn = conectar()
    cur = conn.cursor()

    if papel == "comum" and _ficaria_sem_admin(cur, usuario_id):
        conn.close()
        flash("Não dá para rebaixar o último administrador — o sistema ficaria sem quem administra.", "erro")
        return redirect(url_for("usuarios"))

    cur.execute("UPDATE usuarios SET papel = ? WHERE id = ?", (papel, usuario_id))
    conn.commit()
    conn.close()

    # Se a própria pessoa mudou o próprio papel, a sessão precisa acompanhar.
    if usuario_id == session.get("usuario_id"):
        session["usuario_papel"] = papel

    flash("Papel atualizado.", "sucesso")
    return redirect(url_for("usuarios"))


@app.route("/usuarios/<int:usuario_id>/alternar-acesso", methods=["POST"])
def alternar_acesso_usuario(usuario_id):
    redir = exigir_admin()
    if redir:
        return redir

    if usuario_id == session.get("usuario_id"):
        flash("Você não pode desativar o próprio acesso.", "erro")
        return redirect(url_for("usuarios"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("SELECT ativo FROM usuarios WHERE id = ?", (usuario_id,))
    usuario = cur.fetchone()

    if usuario and usuario["ativo"] and _ficaria_sem_admin(cur, usuario_id):
        conn.close()
        flash("Não dá para desativar o último administrador.", "erro")
        return redirect(url_for("usuarios"))

    cur.execute("UPDATE usuarios SET ativo = 1 - ativo WHERE id = ?", (usuario_id,))
    conn.commit()
    conn.close()

    flash("Acesso atualizado.", "sucesso")
    return redirect(url_for("usuarios"))


@app.route("/usuarios/<int:usuario_id>/senha", methods=["POST"])
def redefinir_senha_usuario(usuario_id):
    redir = exigir_admin()
    if redir:
        return redir

    senha = request.form.get("senha") or ""
    if len(senha) < 6:
        flash("A senha precisa ter pelo menos 6 caracteres.", "erro")
        return redirect(url_for("usuarios"))

    conn = conectar()
    cur = conn.cursor()
    cur.execute(
        "UPDATE usuarios SET senha_hash = ? WHERE id = ?",
        (generate_password_hash(senha), usuario_id),
    )
    conn.commit()
    conn.close()

    flash("Senha redefinida. Peça para a pessoa trocá-la no primeiro acesso.", "sucesso")
    return redirect(url_for("usuarios"))


@app.route("/backups", methods=["GET", "POST"])
def backups():
    redir = exigir_admin()
    if redir:
        return redir

    if request.method == "POST":
        try:
            destino = backup.gerar_backup(motivo="manual")
            flash(f"Backup gerado: {os.path.basename(destino)}", "sucesso")
        except Exception:
            app.logger.exception("Falha ao gerar backup")
            flash("Não foi possível gerar o backup. Veja o log do servidor.", "erro")
        return redirect(url_for("backups"))

    return render_template("backups.html", copias=backup.listar_backups())


@app.route("/backups/<nome>/baixar")
def baixar_backup(nome):
    redir = exigir_admin()
    if redir:
        return redir

    # secure_filename impede que o nome escape da pasta de backups
    # (algo como "../../database.db").
    seguro = secure_filename(nome)
    caminho = os.path.join(backup.PASTA_BACKUPS, seguro)

    if not seguro.endswith(".db") or not os.path.isfile(caminho):
        abort(404)

    return send_file(caminho, as_attachment=True, download_name=seguro)


def _ficaria_sem_admin(cur, usuario_id):
    """True se mexer nesse usuário deixaria o sistema sem nenhum admin ativo."""
    cur.execute(
        "SELECT COUNT(*) AS n FROM usuarios WHERE papel = 'admin' AND ativo = 1 AND id <> ?",
        (usuario_id,),
    )
    return cur.fetchone()["n"] == 0


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@app.route("/dashboard")
def dashboard():
    redir = exigir_login()
    if redir:
        return redir

    ano_atual = request.args.get("ano", type=int) or datetime.date.today().year
    empresa_id = request.args.get("empresa_id", type=int)

    conn = conectar()
    cur = conn.cursor()

    if empresa_id:
        cur.execute("""
            SELECT o.*, e.nome AS empresa_nome
            FROM obras o
            JOIN empresas e ON e.id = o.empresa_id
            WHERE o.empresa_id = ?
            ORDER BY o.status = 'encerrada', o.nome
        """, (empresa_id,))
    else:
        cur.execute("""
            SELECT o.*, e.nome AS empresa_nome
            FROM obras o
            JOIN empresas e ON e.id = o.empresa_id
            ORDER BY o.status = 'encerrada', o.nome
        """)
    obras = [dict(r) for r in cur.fetchall()]

    empresas_lista = listar_empresas(cur)
    anos = anos_com_dados(cur, ano_atual)
    conn.close()

    obra_ids = [o["id"] for o in obras]
    meses = meses_do_ano(ano_atual)
    resultado = calcular_dre_consolidado(obra_ids, meses) if obra_ids else {
        "totais": {c: {campo: 0 for campo in CAMPOS_TOTAIS} for c in meses},
        "acumulado": {campo: 0 for campo in CAMPOS_TOTAIS},
        "por_obra": {},
    }

    acumulado = resultado["acumulado"]

    # Ranking: só entram obras que tiveram movimento no ano. Sem esse filtro, as
    # piores posições ficariam ocupadas por obras zeradas, que não dizem nada.
    ranking = []
    for obra in obras:
        valores = resultado["por_obra"].get(obra["id"])
        if not valores or (not valores["receita_total"] and not valores["custos_total"]):
            continue
        receita = valores["receita_total"]
        ranking.append({
            "id": obra["id"],
            "nome": obra["nome"],
            "codigo": obra["codigo"],
            "receita": receita,
            "lucro_liquido": valores["lucro_liquido"],
            "margem": (valores["lucro_liquido"] / receita * 100) if receita else 0.0,
        })

    ranking.sort(key=lambda o: o["lucro_liquido"], reverse=True)
    melhores = ranking[:6]
    piores = sorted(ranking[-6:], key=lambda o: o["lucro_liquido"])

    # Escala das barrinhas do ranking — proporcional ao maior valor absoluto.
    maior_absoluto = max((abs(o["lucro_liquido"]) for o in ranking), default=0) or 1

    margem_liquida = (acumulado["lucro_liquido"] / acumulado["receita_total"] * 100) \
        if acumulado["receita_total"] else 0.0
    margem_bruta = (acumulado["lucro_bruto"] / acumulado["receita_total"] * 100) \
        if acumulado["receita_total"] else 0.0

    meses_com_dado = [
        c for c in meses
        if resultado["totais"][c]["receita_total"] or resultado["totais"][c]["custos_total"]
    ]

    return render_template(
        "dashboard.html",
        ano_atual=ano_atual,
        anos=anos,
        empresas=empresas_lista,
        empresa_id=empresa_id,
        acumulado=acumulado,
        margem_liquida=margem_liquida,
        margem_bruta=margem_bruta,
        total_obras=len(obras),
        obras_ativas=sum(1 for o in obras if o["status"] != "encerrada"),
        obras_com_movimento=len(ranking),
        meses_com_dado=len(meses_com_dado),
        melhores=melhores,
        piores=piores,
        maior_absoluto=maior_absoluto,
        grafico={
            "labels": [MESES_ABREV[m] for _, m in meses],
            "receita": [round(resultado["totais"][c]["receita_total"], 2) for c in meses],
            "custos": [round(resultado["totais"][c]["custos_total"], 2) for c in meses],
            "liquido": [round(resultado["totais"][c]["lucro_liquido"], 2) for c in meses],
        },
    )


# ---------------------------------------------------------------------------
# Empresas (clientes da contabilidade)
# ---------------------------------------------------------------------------

@app.route("/empresas", methods=["GET", "POST"])
def empresas():
    redir = exigir_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        cnpj = (request.form.get("cnpj") or "").strip() or None

        if nome:
            try:
                cur.execute("INSERT INTO empresas (nome, cnpj) VALUES (?, ?)", (nome, cnpj))
                conn.commit()
                flash("Empresa cadastrada com sucesso.", "sucesso")
            except Exception as e:
                erro_ao_salvar("Não foi possível cadastrar. Verifique se o CNPJ já existe.", e)
        else:
            flash("Informe o nome da empresa.", "erro")

    cur.execute("""
        SELECT e.*, COUNT(o.id) AS total_obras
        FROM empresas e
        LEFT JOIN obras o ON o.empresa_id = e.id
        GROUP BY e.id
        ORDER BY e.nome
    """)
    lista = cur.fetchall()
    conn.close()

    return render_template("empresas.html", empresas=lista)


@app.route("/empresas/<int:empresa_id>/editar", methods=["GET", "POST"])
def editar_empresa(empresa_id):
    redir = exigir_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        cnpj = (request.form.get("cnpj") or "").strip() or None

        if nome:
            try:
                cur.execute("UPDATE empresas SET nome = ?, cnpj = ? WHERE id = ?", (nome, cnpj, empresa_id))
                conn.commit()
                flash("Empresa atualizada com sucesso.", "sucesso")
                conn.close()
                return redirect(url_for("empresas"))
            except Exception as e:
                erro_ao_salvar("Não foi possível salvar. Verifique se o CNPJ já existe em outra empresa.", e)
        else:
            flash("Informe o nome da empresa.", "erro")

    cur.execute("SELECT * FROM empresas WHERE id = ?", (empresa_id,))
    empresa = cur.fetchone()
    conn.close()

    if not empresa:
        flash("Empresa não encontrada.", "erro")
        return redirect(url_for("empresas"))

    return render_template("editar_empresa.html", empresa=empresa)


@app.route("/empresas/<int:empresa_id>/excluir", methods=["POST"])
def excluir_empresa(empresa_id):
    redir = exigir_admin()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) AS total FROM obras WHERE empresa_id = ?", (empresa_id,))
    total_obras = cur.fetchone()["total"]

    if total_obras > 0:
        flash(f"Não é possível excluir: essa empresa ainda tem {total_obras} obra(s) cadastrada(s). Exclua as obras primeiro.", "erro")
    else:
        cur.execute("DELETE FROM empresas WHERE id = ?", (empresa_id,))
        conn.commit()
        flash("Empresa excluída.", "sucesso")

    conn.close()
    return redirect(url_for("empresas"))


# ---------------------------------------------------------------------------
# Obras (centros de custo)
# ---------------------------------------------------------------------------

@app.route("/obras", methods=["GET", "POST"])
def obras():
    redir = exigir_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        codigo = (request.form.get("codigo") or "").strip()
        empresa_id = request.form.get("empresa_id")
        status = request.form.get("status") or "em_andamento"
        data_inicio = request.form.get("data_inicio") or None

        if nome and codigo and empresa_id:
            try:
                cur.execute("""
                    INSERT INTO obras (empresa_id, nome, codigo, status, data_inicio)
                    VALUES (?, ?, ?, ?, ?)
                """, (empresa_id, nome, codigo, status, data_inicio))
                conn.commit()
                flash("Obra cadastrada com sucesso.", "sucesso")
            except Exception as e:
                erro_ao_salvar("Não foi possível cadastrar. Verifique se o código já existe.", e)
        else:
            flash("Preencha nome, código e empresa.", "erro")

    cur.execute("""
        SELECT o.*, e.nome AS empresa_nome
        FROM obras o
        JOIN empresas e ON e.id = o.empresa_id
        ORDER BY o.status = 'encerrada', o.nome
    """)
    lista = cur.fetchall()

    cur.execute("SELECT * FROM empresas ORDER BY nome")
    empresas_lista = cur.fetchall()

    conn.close()

    return render_template("obras.html", obras=lista, empresas=empresas_lista)


@app.route("/obras/<int:obra_id>/editar", methods=["GET", "POST"])
def editar_obra(obra_id):
    redir = exigir_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        codigo = (request.form.get("codigo") or "").strip()
        empresa_id = request.form.get("empresa_id")
        status = request.form.get("status") or "em_andamento"
        data_inicio = request.form.get("data_inicio") or None

        aplica_taxas = 1 if request.form.get("aplica_taxas") == "1" else 0

        if nome and codigo and empresa_id:
            try:
                cur.execute("""
                    UPDATE obras SET nome = ?, codigo = ?, empresa_id = ?, status = ?,
                                     data_inicio = ?, aplica_taxas = ?
                    WHERE id = ?
                """, (nome, codigo, empresa_id, status, data_inicio, aplica_taxas, obra_id))
                conn.commit()
                flash("Obra atualizada com sucesso.", "sucesso")
                conn.close()
                return redirect(url_for("obras"))
            except Exception as e:
                erro_ao_salvar("Não foi possível salvar. Verifique se o código já existe em outra obra.", e)
        else:
            flash("Preencha nome, código e empresa.", "erro")

    cur.execute("SELECT * FROM obras WHERE id = ?", (obra_id,))
    obra = cur.fetchone()
    cur.execute("SELECT * FROM empresas ORDER BY nome")
    empresas_lista = cur.fetchall()
    conn.close()

    if not obra:
        flash("Obra não encontrada.", "erro")
        return redirect(url_for("obras"))

    return render_template("editar_obra.html", obra=obra, empresas=empresas_lista)


@app.route("/obras/<int:obra_id>/excluir", methods=["POST"])
def excluir_obra(obra_id):
    redir = exigir_admin()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    cur.execute("DELETE FROM lancamentos WHERE obra_id = ?", (obra_id,))
    cur.execute("DELETE FROM saldos_anteriores WHERE obra_id = ?", (obra_id,))
    cur.execute("DELETE FROM obras WHERE id = ?", (obra_id,))
    conn.commit()
    conn.close()

    flash("Obra e todos os seus lançamentos foram excluídos.", "sucesso")
    return redirect(url_for("obras"))


@app.route("/obras/<int:obra_id>")
def obra_dre(obra_id):
    redir = exigir_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()
    cur.execute("""
        SELECT o.*, e.nome AS empresa_nome
        FROM obras o JOIN empresas e ON e.id = o.empresa_id
        WHERE o.id = ?
    """, (obra_id,))
    obra = cur.fetchone()

    if not obra:
        conn.close()
        flash("Obra não encontrada.", "erro")
        return redirect(url_for("obras"))

    ano, anos = escolher_ano_da_obra(cur, obra_id, request.args.get("ano", type=int))
    conn.close()

    meses = meses_do_ano(ano)
    resultado = calcular_dre_obra(obra_id, meses, incluir_historico=True)

    # Comparação com o ano anterior — só faz sentido se aquele ano existir de
    # fato para esta obra; senão a variação seria sempre "+100%" contra zero.
    comparativo = None
    if (ano - 1) in anos:
        anterior = calcular_dre_obra(obra_id, meses_do_ano(ano - 1))["acumulado"]
        comparativo = montar_comparativo(resultado["acumulado"], anterior, ano - 1)

    return render_template(
        "obra_dre.html",
        obra=obra,
        ano=ano,
        anos=anos,
        comparativo=comparativo,
        meses=meses,
        meses_nome=MESES_NOME,
        categorias=resultado["categorias"],
        totais=resultado["totais"],
        acumulado=resultado["acumulado"],
        periodos_historicos=resultado["periodos_historicos"],
        totais_historicos=resultado["totais_historicos"],
        total_geral=resultado["total_geral"],
    )


@app.route("/obras/<int:obra_id>/detalhe")
def detalhe_da_celula(obra_id):
    """
    Os lançamentos contábeis por trás de um valor do DRE. Só tem conteúdo para
    competências importadas do Contimatic — os valores que vieram da planilha
    de DRE já chegam somados, sem detalhe por trás.
    """
    redir = exigir_login()
    if redir:
        return redir

    categoria_id = request.args.get("categoria_id", type=int)
    mes = request.args.get("mes", type=int)
    ano = request.args.get("ano", type=int)

    if not (categoria_id and mes and ano):
        abort(400, "Informe categoria, mês e ano.")

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT o.*, e.nome AS empresa_nome
        FROM obras o JOIN empresas e ON e.id = o.empresa_id
        WHERE o.id = ?
    """, (obra_id,))
    obra = cur.fetchone()

    cur.execute("SELECT * FROM categorias_conta WHERE id = ?", (categoria_id,))
    categoria = cur.fetchone()

    if not obra or not categoria:
        conn.close()
        flash("Obra ou categoria não encontrada.", "erro")
        return redirect(url_for("obras"))

    partidas = buscar_partidas(cur, obra_id, categoria_id, ano, mes)

    cur.execute("""
        SELECT valor, origem FROM lancamentos
        WHERE obra_id = ? AND categoria_id = ? AND mes = ? AND ano = ?
    """, (obra_id, categoria_id, mes, ano))
    lancamento = cur.fetchone()
    conn.close()

    return render_template(
        "detalhe_celula.html",
        obra=obra,
        categoria=categoria,
        mes=mes,
        ano=ano,
        partidas=partidas,
        total_partidas=sum(p["valor"] for p in partidas),
        lancamento=lancamento,
    )


@app.route("/obras/<int:obra_id>/exportar")
def exportar_obra_excel(obra_id):
    redir = exigir_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()
    cur.execute("""
        SELECT o.*, e.nome AS empresa_nome
        FROM obras o JOIN empresas e ON e.id = o.empresa_id
        WHERE o.id = ?
    """, (obra_id,))
    obra = cur.fetchone()

    if not obra:
        conn.close()
        flash("Obra não encontrada.", "erro")
        return redirect(url_for("obras"))

    ano, _ = escolher_ano_da_obra(cur, obra_id, request.args.get("ano", type=int))
    conn.close()

    meses = meses_do_ano(ano)
    resultado = calcular_dre_obra(obra_id, meses, incluir_historico=True)

    buffer = gerar_excel_dre_obra(
        obra, ano, meses, MESES_NOME,
        resultado["categorias"], resultado["totais"], resultado["acumulado"],
        periodos_historicos=resultado["periodos_historicos"],
        totais_historicos=resultado["totais_historicos"],
        total_geral=resultado["total_geral"],
    )

    nome_arquivo = f"DRE_{obra['codigo']}_{ano}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Totais consolidados (todas as obras)
# ---------------------------------------------------------------------------

@app.route("/totais")
def totais():
    redir = exigir_login()
    if redir:
        return redir

    ano = request.args.get("ano", type=int) or datetime.date.today().year
    empresa_id = request.args.get("empresa_id", type=int)

    conn = conectar()
    cur = conn.cursor()
    obra_ids = obra_ids_da_empresa(cur, empresa_id)
    empresas_lista = listar_empresas(cur)
    anos = anos_com_dados(cur, ano)
    conn.close()

    meses = meses_do_ano(ano)
    resultado = calcular_dre_consolidado(obra_ids, meses)

    comparativo = None
    if (ano - 1) in anos and obra_ids:
        anterior = calcular_dre_consolidado(obra_ids, meses_do_ano(ano - 1))["acumulado"]
        comparativo = montar_comparativo(resultado["acumulado"], anterior, ano - 1)

    return render_template(
        "totais.html",
        ano=ano,
        anos=anos,
        comparativo=comparativo,
        empresas=empresas_lista,
        empresa_id=empresa_id,
        total_obras=len(obra_ids),
        meses=meses,
        meses_nome=MESES_NOME,
        totais=resultado["totais"],
        acumulado=resultado["acumulado"],
    )


@app.route("/totais/exportar")
def exportar_totais_excel():
    redir = exigir_login()
    if redir:
        return redir

    ano = request.args.get("ano", type=int) or datetime.date.today().year
    empresa_id = request.args.get("empresa_id", type=int)

    conn = conectar()
    cur = conn.cursor()
    obra_ids = obra_ids_da_empresa(cur, empresa_id)

    sufixo = ""
    if empresa_id:
        cur.execute("SELECT nome FROM empresas WHERE id = ?", (empresa_id,))
        empresa = cur.fetchone()
        if empresa:
            sufixo = "_" + re.sub(r"[^A-Za-z0-9]+", "_", empresa["nome"]).strip("_")[:30]
    conn.close()

    meses = meses_do_ano(ano)
    resultado = calcular_dre_consolidado(obra_ids, meses)

    buffer = gerar_excel_dre_consolidado(ano, meses, MESES_NOME, resultado["totais"], resultado["acumulado"])

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"DRE_Consolidado{sufixo}_{ano}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Plano de contas (categorias)
# ---------------------------------------------------------------------------

@app.route("/categorias", methods=["GET", "POST"])
def categorias():
    redir = exigir_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        codigo = (request.form.get("codigo") or "").strip() or None
        tipo = request.form.get("tipo")

        if nome and tipo in ("receita", "custo"):
            cur.execute("SELECT COALESCE(MAX(ordem), 0) + 1 AS prox FROM categorias_conta")
            ordem = cur.fetchone()["prox"]
            try:
                cur.execute("""
                    INSERT INTO categorias_conta (codigo, nome, tipo, ordem, ativo)
                    VALUES (?, ?, ?, ?, 1)
                """, (codigo, nome, tipo, ordem))
                conn.commit()
                flash("Categoria adicionada ao plano de contas.", "sucesso")
            except Exception as e:
                erro_ao_salvar("Já existe uma categoria com esse nome.", e)
        else:
            flash("Informe nome e tipo (receita ou custo).", "erro")

    # Uso real de cada conta: a tela de revisão precisa mostrar o que pode ser
    # mexido sem risco (conta zerada) e o que já carrega histórico.
    cur.execute("""
        SELECT c.*,
               (SELECT COUNT(*) FROM lancamentos l
                 WHERE l.categoria_id = c.id AND l.valor <> 0) AS usos,
               (SELECT COUNT(*) FROM saldos_anteriores s
                 WHERE s.categoria_id = c.id) AS usos_historicos
        FROM categorias_conta c
        ORDER BY c.tipo DESC, c.ordem
    """)
    lista = [dict(r) for r in cur.fetchall()]
    conn.close()

    return render_template(
        "categorias.html",
        categorias=lista,
        duplicatas=_sugerir_duplicatas(lista),
    )


def _sugerir_duplicatas(categorias, limiar=0.82):
    """
    Aponta pares de nomes muito parecidos, do mesmo tipo.

    A importação reconhece variações de acento e caixa, mas não abreviação:
    "Seguro Riscos Execução de Serv. Trab." e "Seguro Riscos Execução de
    Serviços" entram como duas contas e dividem o valor entre duas linhas do DRE.
    """
    import difflib

    def chave(nome):
        return re.sub(r"\s+", " ", _sem_acentos(nome)).strip().lower()

    pares = []
    for i, a in enumerate(categorias):
        for b in categorias[i + 1:]:
            if a["tipo"] != b["tipo"]:
                continue
            semelhanca = difflib.SequenceMatcher(None, chave(a["nome"]), chave(b["nome"])).ratio()
            if semelhanca >= limiar:
                # Sugere manter a que já tem histórico.
                manter, mesclar = (a, b) if (a["usos"] or 0) >= (b["usos"] or 0) else (b, a)
                pares.append({"manter": manter, "mesclar": mesclar, "semelhanca": semelhanca})

    return sorted(pares, key=lambda p: -p["semelhanca"])


def _sem_acentos(texto):
    import unicodedata

    nfkd = unicodedata.normalize("NFKD", str(texto or ""))
    return "".join(c for c in nfkd if not unicodedata.combining(c))


@app.route("/categorias/<int:categoria_id>/editar", methods=["POST"])
def editar_categoria(categoria_id):
    redir = exigir_login()
    if redir:
        return redir

    nome = (request.form.get("nome") or "").strip()
    codigo = (request.form.get("codigo") or "").strip() or None
    tipo = request.form.get("tipo")

    if not nome or tipo not in ("receita", "custo"):
        flash("Informe nome e tipo (receita ou custo).", "erro")
        return redirect(url_for("categorias"))

    conn = conectar()
    cur = conn.cursor()

    # Trocar o tipo de uma conta que já tem lançamento move dinheiro de receita
    # para custo (ou o contrário) em todo o histórico, em todas as obras.
    cur.execute("SELECT tipo FROM categorias_conta WHERE id = ?", (categoria_id,))
    atual = cur.fetchone()
    if atual and atual["tipo"] != tipo:
        cur.execute(
            "SELECT COUNT(*) AS n FROM lancamentos WHERE categoria_id = ? AND valor <> 0",
            (categoria_id,),
        )
        if cur.fetchone()["n"] and not usuario_eh_admin():
            conn.close()
            flash(
                "Trocar receita/custo de uma conta que já tem lançamento altera o DRE de "
                "todas as obras — só a administradora pode fazer isso.",
                "erro",
            )
            return redirect(url_for("categorias"))

    try:
        cur.execute(
            "UPDATE categorias_conta SET nome = ?, codigo = ?, tipo = ?, origem = 'manual' WHERE id = ?",
            (nome, codigo, tipo, categoria_id),
        )
        conn.commit()
        flash("Categoria atualizada.", "sucesso")
    except Exception as e:
        erro_ao_salvar("Já existe outra categoria com esse nome.", e)

    conn.close()
    return redirect(url_for("categorias"))


@app.route("/categorias/mesclar", methods=["POST"])
def mesclar_categorias():
    """
    Junta duas contas numa só: todo o histórico da origem passa para o destino
    e a origem é removida. Não tem desfazer, por isso é da administradora.
    """
    redir = exigir_admin()
    if redir:
        return redir

    origem_id = request.form.get("origem_id", type=int)
    destino_id = request.form.get("destino_id", type=int)

    if not origem_id or not destino_id or origem_id == destino_id:
        flash("Escolha duas categorias diferentes para mesclar.", "erro")
        return redirect(url_for("categorias"))

    conn = conectar()
    cur = conn.cursor()

    cur.execute("SELECT id, nome, tipo FROM categorias_conta WHERE id IN (?, ?)", (origem_id, destino_id))
    encontradas = {r["id"]: r for r in cur.fetchall()}

    if len(encontradas) != 2:
        conn.close()
        flash("Categoria não encontrada.", "erro")
        return redirect(url_for("categorias"))

    origem, destino = encontradas[origem_id], encontradas[destino_id]

    if origem["tipo"] != destino["tipo"]:
        conn.close()
        flash("Não dá para mesclar uma receita com um custo.", "erro")
        return redirect(url_for("categorias"))

    try:
        movidos = _mesclar_categoria(cur, origem_id, destino_id)
        cur.execute("DELETE FROM categorias_conta WHERE id = ?", (origem_id,))
        conn.commit()
        flash(
            f'"{origem["nome"]}" foi mesclada em "{destino["nome"]}". '
            f'{movidos} valor(es) passaram para a conta de destino.',
            "sucesso",
        )
    except Exception:
        conn.rollback()
        app.logger.exception("Falha ao mesclar categorias")
        flash("Erro ao mesclar. Nada foi alterado.", "erro")

    conn.close()
    return redirect(url_for("categorias"))


def _mesclar_categoria(cur, origem_id, destino_id):
    """
    Move lançamentos, períodos históricos e partidas da origem para o destino.

    Onde o destino já tem valor na mesma competência, os dois são somados — sem
    isso a restrição de unicidade barraria a mesclagem e o valor da origem se
    perderia em silêncio.
    """
    movidos = 0

    # Lançamentos mensais
    cur.execute("SELECT * FROM lancamentos WHERE categoria_id = ?", (origem_id,))
    for linha in cur.fetchall():
        cur.execute(
            """
            SELECT id, valor FROM lancamentos
            WHERE obra_id = ? AND categoria_id = ? AND mes = ? AND ano = ?
            """,
            (linha["obra_id"], destino_id, linha["mes"], linha["ano"]),
        )
        existente = cur.fetchone()

        if existente:
            cur.execute(
                "UPDATE lancamentos SET valor = ? WHERE id = ?",
                (existente["valor"] + linha["valor"], existente["id"]),
            )
            cur.execute("DELETE FROM lancamentos WHERE id = ?", (linha["id"],))
        else:
            cur.execute(
                "UPDATE lancamentos SET categoria_id = ? WHERE id = ?", (destino_id, linha["id"])
            )
        movidos += 1

    # Períodos históricos agregados
    cur.execute("SELECT * FROM saldos_anteriores WHERE categoria_id = ?", (origem_id,))
    for linha in cur.fetchall():
        cur.execute(
            """
            SELECT id, valor FROM saldos_anteriores
            WHERE obra_id = ? AND categoria_id = ? AND periodo_descricao = ?
            """,
            (linha["obra_id"], destino_id, linha["periodo_descricao"]),
        )
        existente = cur.fetchone()

        if existente:
            cur.execute(
                "UPDATE saldos_anteriores SET valor = ? WHERE id = ?",
                (existente["valor"] + linha["valor"], existente["id"]),
            )
            cur.execute("DELETE FROM saldos_anteriores WHERE id = ?", (linha["id"],))
        else:
            cur.execute(
                "UPDATE saldos_anteriores SET categoria_id = ? WHERE id = ?",
                (destino_id, linha["id"]),
            )
        movidos += 1

    # Detalhe contábil (Contimatic) e o de-para de contas
    cur.execute("UPDATE partidas SET categoria_id = ? WHERE categoria_id = ?", (destino_id, origem_id))
    cur.execute("UPDATE contas_map SET categoria_id = ? WHERE categoria_id = ?", (destino_id, origem_id))

    return movidos


# POST, não GET: era a única rota que alterava dado por GET. Um link assim é
# disparado por pré-carregamento do navegador, antivírus ou indexador — bastava
# alguém passar o mouse por cima para desativar uma conta do plano.
@app.route("/categorias/<int:categoria_id>/alternar-status", methods=["POST"])
def alternar_status_categoria(categoria_id):
    redir = exigir_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()
    cur.execute("UPDATE categorias_conta SET ativo = 1 - ativo WHERE id = ?", (categoria_id,))
    cur.execute("SELECT nome, ativo FROM categorias_conta WHERE id = ?", (categoria_id,))
    categoria = cur.fetchone()
    conn.commit()
    conn.close()

    if categoria:
        estado = "ativada" if categoria["ativo"] else "desativada"
        flash(f'Categoria "{categoria["nome"]}" {estado}.', "sucesso")

    return redirect(url_for("categorias"))


# ---------------------------------------------------------------------------
# Taxas configuráveis (impostos, IRPJ/CSLL, adm., financeiras)
# ---------------------------------------------------------------------------

@app.route("/taxas", methods=["GET", "POST"])
def taxas():
    # Consultar as taxas é livre — todo mundo precisa saber com que alíquota o
    # DRE foi calculado. Criar uma vigência nova é da dona: muda o resultado de
    # todas as obras de uma vez.
    redir = exigir_login()
    if redir:
        return redir

    if request.method == "POST":
        bloqueio = exigir_admin()
        if bloqueio:
            return bloqueio

    conn = conectar()
    cur = conn.cursor()

    if request.method == "POST":
        chave = request.form.get("chave")
        percentual = request.form.get("percentual")
        vigencia_inicio = request.form.get("vigencia_inicio")

        if chave and percentual and vigencia_inicio:
            cur.execute("""
                SELECT descricao, base_calculo FROM taxas
                WHERE chave = ? ORDER BY vigencia_inicio DESC LIMIT 1
            """, (chave,))
            referencia = cur.fetchone()

            if referencia:
                ano, mes = vigencia_inicio.split("-")
                ano, mes = int(ano), int(mes)
                mes_anterior = mes - 1
                ano_fim = ano
                if mes_anterior == 0:
                    mes_anterior = 12
                    ano_fim -= 1
                fim_anterior = f"{ano_fim:04d}-{mes_anterior:02d}"

                cur.execute("""
                    UPDATE taxas SET vigencia_fim = ?
                    WHERE chave = ? AND vigencia_fim IS NULL
                """, (fim_anterior, chave))

                cur.execute("""
                    INSERT INTO taxas (chave, descricao, percentual, base_calculo, vigencia_inicio, vigencia_fim)
                    VALUES (?, ?, ?, ?, ?, NULL)
                """, (chave, referencia["descricao"], float(percentual), referencia["base_calculo"], vigencia_inicio))

                conn.commit()
                flash("Nova vigência de taxa cadastrada.", "sucesso")
            else:
                flash("Taxa não encontrada.", "erro")
        else:
            flash("Preencha todos os campos.", "erro")

    cur.execute("SELECT * FROM taxas ORDER BY chave, vigencia_inicio DESC")
    lista = cur.fetchall()
    conn.close()

    return render_template("taxas.html", taxas=lista)


# ---------------------------------------------------------------------------
# Lançamentos manuais (grade categoria x valor, para uma obra/mês/ano)
# ---------------------------------------------------------------------------

@app.route("/lancamentos", methods=["GET", "POST"])
def lancamentos_manual():
    redir = exigir_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    cur.execute("SELECT id, nome, codigo FROM obras ORDER BY nome")
    obras_lista = cur.fetchall()

    obra_id = request.args.get("obra_id", type=int) or request.form.get("obra_id", type=int)
    mes = request.args.get("mes", type=int) or request.form.get("mes", type=int)
    ano = request.args.get("ano", type=int) or request.form.get("ano", type=int)

    if request.method == "POST" and obra_id and mes and ano:
        cur.execute("SELECT id FROM categorias_conta WHERE ativo = 1")
        categoria_ids = [r["id"] for r in cur.fetchall()]

        import datetime
        agora = datetime.datetime.now().isoformat()

        for categoria_id in categoria_ids:
            valor = parse_valor_br(request.form.get(f"valor_{categoria_id}"))

            cur.execute("""
                INSERT INTO lancamentos (obra_id, categoria_id, mes, ano, valor, origem, atualizado_em)
                VALUES (?, ?, ?, ?, ?, 'manual', ?)
                ON CONFLICT (obra_id, categoria_id, mes, ano)
                DO UPDATE SET valor = excluded.valor, origem = 'manual', atualizado_em = excluded.atualizado_em
            """, (obra_id, categoria_id, mes, ano, valor, agora))

        conn.commit()
        flash("Lançamentos salvos com sucesso.", "sucesso")

    categorias_com_valor = []
    if obra_id and mes and ano:
        cur.execute("""
            SELECT c.id, c.codigo, c.nome, c.tipo,
                   COALESCE(l.valor, 0) AS valor
            FROM categorias_conta c
            LEFT JOIN lancamentos l
                ON l.categoria_id = c.id AND l.obra_id = ? AND l.mes = ? AND l.ano = ?
            WHERE c.ativo = 1
            ORDER BY c.tipo DESC, c.ordem
        """, (obra_id, mes, ano))
        categorias_com_valor = cur.fetchall()

    conn.close()

    return render_template(
        "lancamentos.html",
        obras=obras_lista,
        obra_id=obra_id,
        mes=mes,
        ano=ano,
        meses_nome=MESES_NOME,
        categorias=categorias_com_valor,
    )


# ---------------------------------------------------------------------------
# Importação de dados via planilha
# ---------------------------------------------------------------------------

@app.route("/importar-dados", methods=["GET", "POST"])
def importar_dados():
    redir = exigir_login()
    if redir:
        return redir

    if request.method == "POST":
        arquivo = request.files.get("arquivo")

        if not arquivo or arquivo.filename == "":
            flash("Selecione um arquivo Excel para importar.", "erro")
            return redirect(url_for("importar_dados"))

        if not arquivo_permitido(arquivo.filename):
            flash("Formato inválido. Envie apenas arquivos .xlsx ou .xls.", "erro")
            return redirect(url_for("importar_dados"))

        filename = secure_filename(arquivo.filename)
        caminho_arquivo = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        arquivo.save(caminho_arquivo)

        nomes_abas = []
        try:
            wb_check = openpyxl.load_workbook(caminho_arquivo, read_only=True)
            nomes_abas = wb_check.sheetnames
            wb_check.close()
        except Exception:
            pass

        eh_arquivo_dre_real = any(
            re.search(r"OBRA\s*\d+", nome.upper()) or "DEMAIS OBRAS" in nome.upper()
            for nome in nomes_abas
        )

        # Marcado na tela: por padrão a importação preserva o que foi lançado à mão.
        sobrescrever_manuais = request.form.get("sobrescrever_manuais") == "1"

        if eh_arquivo_dre_real:
            return _importar_arquivo_dre_real(caminho_arquivo, filename, sobrescrever_manuais)

        try:
            df = pd.read_excel(caminho_arquivo)
            df.columns = [str(col).strip().lower() for col in df.columns]

            colunas_necessarias = ["codigo_obra", "categoria", "mes", "ano", "valor"]
            faltando = [c for c in colunas_necessarias if c not in df.columns]
            if faltando:
                flash("A planilha está faltando estas colunas: " + ", ".join(faltando), "erro")
                return redirect(url_for("importar_dados"))

            conn = conectar()
            cur = conn.cursor()

            cur.execute("SELECT id, codigo FROM obras")
            obra_por_codigo = {r["codigo"]: r["id"] for r in cur.fetchall()}

            cur.execute("SELECT id, codigo, nome FROM categorias_conta")
            todas_categorias = cur.fetchall()
            categoria_por_codigo = {r["codigo"]: r["id"] for r in todas_categorias if r["codigo"]}
            categoria_por_nome = {r["nome"].strip().lower(): r["id"] for r in todas_categorias}

            agora = datetime.datetime.now().isoformat()

            importacao_id = _criar_registro_importacao(cur, filename, None)
            registro = RegistroImportacao(cur, importacao_id, sobrescrever_manuais)

            total_ok = 0
            total_ignorados = 0

            for _, linha in df.iterrows():
                codigo_obra = str(linha.get("codigo_obra", "")).strip()
                categoria_ref = str(linha.get("categoria", "")).strip()

                obra_id = obra_por_codigo.get(codigo_obra)
                categoria_id = categoria_por_codigo.get(categoria_ref) or categoria_por_nome.get(categoria_ref.lower())

                if not obra_id or not categoria_id:
                    total_ignorados += 1
                    continue

                try:
                    mes = int(float(linha["mes"]))
                    ano = int(float(linha["ano"]))
                    valor_bruto = linha["valor"]
                    if isinstance(valor_bruto, str):
                        valor = float(valor_bruto.replace("R$", "").replace(".", "").replace(",", ".").strip())
                    else:
                        valor = float(valor_bruto)
                    valor = abs(valor)
                except (ValueError, TypeError):
                    total_ignorados += 1
                    continue

                if mes < 1 or mes > 12:
                    total_ignorados += 1
                    continue

                if registro.gravar_lancamento(obra_id, categoria_id, mes, ano, valor, agora):
                    total_ok += 1

            _fechar_registro_importacao(cur, importacao_id, registro)
            conn.commit()
            conflitos = _descrever_conflitos(cur, registro)
            conn.close()

            msg = f"Importação concluída. {total_ok} lançamento(s) importado(s)."
            if total_ignorados:
                msg += f" {total_ignorados} linha(s) ignorada(s) (obra ou categoria não encontrada)."
            flash(msg, "sucesso")
            _avisar_conflitos(registro, conflitos)
            return redirect(url_for("importar_dados"))

        except Exception as e:
            app.logger.exception("Falha ao processar planilha simples")
            flash(f"Erro ao processar a planilha: {str(e)}", "erro")
            return redirect(url_for("importar_dados"))

    conn = conectar()
    cur = conn.cursor()
    cur.execute("""
        SELECT i.*, e.nome AS empresa_nome, u.nome AS usuario_nome
        FROM importacoes i
        LEFT JOIN empresas e ON e.id = i.empresa_id
        LEFT JOIN usuarios u ON u.id = i.usuario_id
        ORDER BY i.id DESC LIMIT 10
    """)
    historico = [dict(r) for r in cur.fetchall()]
    conn.close()

    # Só a importação mais recente ainda ativa pode ser desfeita: reverter uma
    # antiga por cima de outra mais nova ressuscitaria valores obsoletos.
    id_reversivel = next((h["id"] for h in historico if not h["desfeita_em"]), None)

    return render_template("importar_dados.html", historico=historico, id_reversivel=id_reversivel)


def _detectar_nome_empresa(caminho_arquivo, nomes_abas):
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    for nome_aba in nomes_abas:
        if re.search(r"OBRA\s*\d+", nome_aba.upper()) or "DEMAIS OBRAS" in nome_aba.upper():
            ws = wb[nome_aba]
            for r in range(1, 4):
                for c in (1, 2):
                    valor = ws.cell(row=r, column=c).value
                    if valor and len(str(valor).strip()) > 4 and "DEMONSTRAÇÃO" not in str(valor).upper():
                        return str(valor).strip()
    return "Empresa Importada"


def _obter_ou_criar_empresa(cur, nome_empresa):
    cur.execute("SELECT id FROM empresas WHERE nome = ?", (nome_empresa,))
    row = cur.fetchone()
    if row:
        return row["id"]
    cur.execute("INSERT INTO empresas (nome, cnpj) VALUES (?, NULL)", (nome_empresa,))
    return cur.lastrowid


def _criar_registro_importacao(cur, arquivo, empresa_id):
    cur.execute(
        """
        INSERT INTO importacoes (arquivo, empresa_id, usuario_id, criado_em)
        VALUES (?, ?, ?, ?)
        """,
        (arquivo, empresa_id, session.get("usuario_id"), datetime.datetime.now().isoformat()),
    )
    return cur.lastrowid


def _fechar_registro_importacao(cur, importacao_id, registro):
    cur.execute(
        """
        UPDATE importacoes
        SET lancamentos_gravados = ?, saldos_gravados = ?,
            manuais_preservados = ?, manuais_sobrescritos = ?
        WHERE id = ?
        """,
        (registro.lancamentos_gravados, registro.saldos_gravados,
         registro.manuais_preservados, registro.manuais_sobrescritos, importacao_id),
    )


def _descrever_conflitos(cur, registro):
    """Troca os ids da amostra de conflitos por nomes, para a mensagem na tela."""
    descritos = []
    for c in registro.conflitos:
        cur.execute("SELECT nome FROM obras WHERE id = ?", (c["obra_id"],))
        obra = cur.fetchone()
        cur.execute("SELECT nome FROM categorias_conta WHERE id = ?", (c["categoria_id"],))
        categoria = cur.fetchone()
        descritos.append(
            f"{obra['nome'] if obra else c['obra_id']} · "
            f"{categoria['nome'] if categoria else c['categoria_id']} · "
            f"{MESES_ABREV[c['mes']]}/{c['ano']}: "
            f"manteve {filtro_moeda(c['valor_manual'])} "
            f"(planilha trazia {filtro_moeda(c['valor_planilha'])})"
        )
    return descritos


def _avisar_conflitos(registro, conflitos):
    if registro.manuais_preservados:
        flash(
            f"{registro.manuais_preservados} valor(es) lançado(s) manualmente foram "
            f"PRESERVADOS e não vieram da planilha. Para que a planilha prevaleça, "
            f"reenvie marcando a opção de sobrescrever lançamentos manuais.",
            "erro",
        )
        for descricao in conflitos:
            flash(f"Preservado — {descricao}", "erro")

    if registro.manuais_sobrescritos:
        flash(
            f"{registro.manuais_sobrescritos} valor(es) lançado(s) manualmente foram "
            f"substituídos pelos da planilha, conforme solicitado. Use 'Desfazer' "
            f"abaixo se não era o esperado.",
            "erro",
        )


def _importar_arquivo_dre_real(caminho_arquivo, nome_arquivo, sobrescrever_manuais):
    conn = conectar()
    cur = conn.cursor()

    wb_check = openpyxl.load_workbook(caminho_arquivo, read_only=True)
    nomes_abas = wb_check.sheetnames
    wb_check.close()

    nome_empresa = _detectar_nome_empresa(caminho_arquivo, nomes_abas)
    empresa_id = _obter_ou_criar_empresa(cur, nome_empresa)
    conn.commit()

    importacao_id = _criar_registro_importacao(cur, nome_arquivo, empresa_id)
    registro = RegistroImportacao(cur, importacao_id, sobrescrever_manuais)

    try:
        resumo = importar_planilha_dre_real(caminho_arquivo, cur, empresa_id, registro)
        _fechar_registro_importacao(cur, importacao_id, registro)
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        app.logger.exception("Falha ao processar arquivo de DRE")
        flash(f"Erro ao processar o arquivo: {str(e)}", "erro")
        return redirect(url_for("importar_dados"))

    conflitos = _descrever_conflitos(cur, registro)
    conn.close()

    partes = [
        f"Empresa: {nome_empresa}.",
        f"{len(resumo['abas_processadas'])} obra(s) lida(s).",
        f"{resumo['lancamentos_gravados']} lançamento(s) gravado(s).",
    ]
    if resumo["obras_criadas"]:
        partes.append(f"{len(resumo['obras_criadas'])} obra(s) nova(s) cadastrada(s) automaticamente.")
    if resumo["categorias_criadas"]:
        partes.append(f"{len(resumo['categorias_criadas'])} categoria(s) nova(s) no plano de contas.")
    if resumo["saldos_anteriores_gravados"]:
        partes.append(f"{resumo['saldos_anteriores_gravados']} valor(es) de período histórico (sem mês definido) também gravado(s) — veja na tela de cada obra.")
    if resumo["abas_ignoradas"]:
        partes.append(f"{len(resumo['abas_ignoradas'])} aba(s) ignorada(s) (veja detalhes abaixo).")

    flash(" ".join(partes), "sucesso")
    _avisar_conflitos(registro, conflitos)

    for nome_aba, motivo in resumo["abas_ignoradas"]:
        flash(f"Aba '{nome_aba}' ignorada: {motivo}", "erro")

    return redirect(url_for("importar_dados"))


@app.route("/importacoes/<int:importacao_id>/desfazer", methods=["POST"])
def desfazer_importacao_rota(importacao_id):
    redir = exigir_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    cur.execute("SELECT id FROM importacoes WHERE desfeita_em IS NULL ORDER BY id DESC LIMIT 1")
    mais_recente = cur.fetchone()

    if not mais_recente or mais_recente["id"] != importacao_id:
        conn.close()
        flash(
            "Só é possível desfazer a importação mais recente. Desfazer uma antiga "
            "por cima de outra mais nova traria de volta valores já superados.",
            "erro",
        )
        return redirect(url_for("importar_dados"))

    try:
        resultado = desfazer_importacao(cur, importacao_id)
        conn.commit()
    except ValueError as e:
        conn.rollback()
        conn.close()
        flash(str(e), "erro")
        return redirect(url_for("importar_dados"))
    except Exception:
        conn.rollback()
        conn.close()
        app.logger.exception("Falha ao desfazer importação")
        flash("Erro inesperado ao desfazer a importação. Nada foi alterado.", "erro")
        return redirect(url_for("importar_dados"))

    conn.close()

    flash(
        f"Importação desfeita. {resultado['restaurados']} valor(es) voltaram ao que eram antes "
        f"e {resultado['removidos']} valor(es) criados por ela foram removidos.",
        "sucesso",
    )
    return redirect(url_for("importar_dados"))


if __name__ == "__main__":
    # debug=True expõe um console Python remoto a quem alcançar a porta.
    # Fica desligado por padrão; ligue com INOV_DEBUG=1 durante o desenvolvimento.
    app.run(debug=os.environ.get("INOV_DEBUG") == "1")
